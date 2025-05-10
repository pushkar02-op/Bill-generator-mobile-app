from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.invoice_tracker import get_next_invoice_number

invoice_number = get_next_invoice_number()

def generate_excel_bill(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # === HEADER (Green section) ===
    for row_num, value in enumerate(["TAX INVOICE", "A.G AGRO", "TEGHRA,BEGUSARAI-851133", "10HVGPD2399M1ZC"], start=1):
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=12)
        cell = ws.cell(row=row_num, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        if row_num == 1:
            ws.row_dimensions[row_num].height = 36  
            cell.font = Font(size=15, bold=True, italic=True)
        else:
            ws.row_dimensions[row_num].height = 14
            cell.font = Font(size=12, bold=True, italic=True)

    # === SECTION HEADERS ===
    ws.merge_cells("A5:D5")
    ws.merge_cells("E5:H5")
    ws.merge_cells("I5:L5")
    section_headers = {"A5": "BILL TO", "E5": "PLACE OF SUPPLY", "I5": "BILL DETAILS:"}

    for cell_ref, text in section_headers.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.font = Font(bold=True, italic=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 12

    # === SECTION CONTENT (Blue) ===
    ws.merge_cells("A6:D7")
    ws.merge_cells("E6:H7")
    ws.merge_cells("I6:L7")

    ws["A6"] = data["bill_to"]
    ws["E6"] = data["place_of_supply"]
    ws["I6"] = (
        f"INVOICE NO: {invoice_number}\n"
        f"DELIVERY DATE: {data['delivery_date']}\n"
        f"VENDOR CODE: {data['vendor_code']}\n"
        f"SITE CODE: {data['site_code']}"
    )

    for col in range(1, 13):
        for row in range(6, 8):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for cell in ["A6", "E6", "I6"]:
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cell].font = Font(size=9, bold=True, italic=True)

    for row in range(6, 8):
        ws.row_dimensions[row].height = 36

    # === GST and PO ===
    ws.merge_cells("A8:F8")
    ws.merge_cells("G8:L8")
    ws["A8"] = f"GST: {data['GST']}"
    ws["G8"] = f"PO-{data['PO']}"
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A8"].font = Font(bold=True, italic=True)
    ws["G8"].font = Font(bold=True, italic=True)
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws["A8"].fill = blue_fill
    ws["G8"].fill = blue_fill

    ws.row_dimensions[8].height = 12

    # === TABLE HEADERS ===
    headers = [
        "ARTICLE CODE", "HSN CODE", "Article Description", "Grammage", "Quantity",
        "Rate", "Taxable Value", "SGST Rate", "SGST Amount",
        "CGST Rate", "CGST Amount", "Total Amount"
    ]
    header_row = 9
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, italic=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
    ws.row_dimensions[header_row].height = 30

    # === DATA ROWS ===
    data_start_row = header_row + 1
    for row_offset, item in enumerate(data["items"]):
        for col_idx, value in enumerate(item, start=1):
            cell = ws.cell(row=data_start_row + row_offset, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_idx in (1, 2):
                cell.number_format = '0'
            elif col_idx in (5, 6, 7, 9, 11, 12):
                cell.number_format = '#,##0.00'
            elif col_idx in (8, 10):
                cell.number_format = '0.00%'
            else:
                cell.number_format = '@'

            if (data_start_row + row_offset) % 2 == 0:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )

    # === FOOTER (Sub Total) ===
    footer_row = data_start_row + len(data["items"])
    label_cell = ws.cell(row=footer_row, column=1, value="Sub Total")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )

    sum_cols = [5, 7, 8, 9, 10, 11, 12]
    for col_idx in sum_cols:
        col_letter = get_column_letter(col_idx)
        sum_cell = ws.cell(
            row=footer_row,
            column=col_idx,
            value=f"=SUM({col_letter}{data_start_row}:{col_letter}{footer_row-1})"
        )
        if col_idx in (8, 10):
            sum_cell.number_format = '0.00%'
        else:
            sum_cell.number_format = '#,##0.00'
        sum_cell.font = Font(bold=True)
        sum_cell.alignment = Alignment(horizontal="center", vertical="center")
        sum_cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )

    # === COLUMN WIDTHS ===
    min_width = 10
    max_width = 40
    max_row = footer_row + 1

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col - 1])
        for row in range(data_start_row, max_row):
            value = ws.cell(row=row, column=col).value
            if value:
                max_length = max(max_length, len(str(value)))
        adjusted_width = min(max(max_length, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

        # === ADDITIONAL FOOTER INFO ===
    current_row = footer_row + 1

    # Row: Misc. Charges
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    misc_cell = ws.cell(row=current_row, column=1, value="Misc. Charges (Including Freight, Octroi, Loading, Unloading, etc.)")
    misc_cell.font = Font(bold=True, size=9, italic=True)
    misc_cell.alignment = Alignment(horizontal="right", vertical="center")
    current_row += 1

    # Row: Grand Total
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    grand_cell = ws.cell(row=current_row, column=1, value="Grand Total (Rounded Off)")
    grand_cell.font = Font(bold=True, size=9 ,italic=True)
    grand_cell.alignment = Alignment(horizontal="right", vertical="center")
    current_row += 1

    # Blank Row
    current_row += 1

    # Thank You Row (Center-aligned)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    thank_you_cell = ws.cell(row=current_row, column=1, value="THANK YOU FOR YOUR BUSINESS")
    thank_you_cell.font = Font(bold=True, size=8, italic=True)
    thank_you_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Clear borders for entire "THANK YOU" row (including the merged cell)
    no_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None),
    )

    for col in range(1, 13):
        cell = ws.cell(row=current_row, column=col)
        cell.border = no_border

    current_row += 1

    # Clear borders for 3 rows after "THANK YOU"
    for row in range(current_row, current_row + 3):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = no_border

    # Signature at bottom-left
    signature_cell = ws.cell(row=current_row + 3, column=1, value="Signature")
    signature_cell.font = Font(bold=True, size=8)
    signature_cell.alignment = Alignment(horizontal="left", vertical="center")

    # === SAVE FILE ===
    wb.save(filename)
    print(f"✅ Bill saved to {filename}")
