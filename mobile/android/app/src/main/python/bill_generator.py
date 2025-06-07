import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pdf_generator import generate_pdf_bill
from openpyxl.worksheet.page import PageMargins

BASE_DIR      = Path(__file__).parent

COUNTER_FILE  = Path("/sdcard/Documents/bills") / "invoice_counter.json"
METADATA_FILE = BASE_DIR / "metadata.json"

def _get_next_invoice_number() -> int:
    if not COUNTER_FILE.exists():
        COUNTER_FILE.write_text(json.dumps({"last_invoice": 1150}))
    
    data = json.loads(COUNTER_FILE.read_text())
    next_inv = data["last_invoice"] + 1
    COUNTER_FILE.write_text(json.dumps({"last_invoice": next_inv}))
    return next_inv


def _extract_po_and_date(filename: str) :
    """
    Extract PO number and delivery date from a structured filename.
    Example: '15467510000244_20250422_055526.xlsx' -> ('15467510000244', '22-04-2025')
    """
    stem = Path(filename).stem
    parts = stem.split("_", maxsplit=2)
    po = parts[0] if parts else "N/A"
    raw = parts[1] if len(parts) > 1 else ""
    try:
        dt = datetime.strptime(raw, "%Y%m%d")
        return po, dt.strftime("%d-%m-%Y")
    except:
        return po, raw or "N/A"

def _transform_items(df: pd.DataFrame) -> list:
    df = df.copy()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce", downcast="integer").fillna(0).astype(int)
    df["Landing Rate"] = pd.to_numeric(df["Landing Rate"], errors="coerce", downcast="float").fillna(0.0).astype(float)
    df["Item Code"]    = pd.to_numeric(df["Item Code"],    errors="coerce", downcast="integer").fillna(0).astype(int)
    df["HSN Code"]     = pd.to_numeric(df["HSN Code"],     errors="coerce", downcast="integer").fillna(0).astype(int)
    df["Product Description"] = df["Product Description"].astype(str).fillna("")
    df["Grammage"]              = df["Grammage"].astype(str).fillna("")
    df = df.dropna(subset=["Quantity","Landing Rate","Item Code","HSN Code","Product Description","Grammage"])
    
    items = []
    for _, row in df.iterrows():
        qty, rate = row["Quantity"], row["Landing Rate"]
        total = round(qty * rate, 2)
        items.append([
            row["Item Code"], row["HSN Code"],
            row["Product Description"].strip(), row["Grammage"].strip(),
            qty, rate, total,
            0.00, 0.0, 0.00, 0.0, total
        ])
    return items[:-3] if len(items) >= 3 else items

def generate_bill(input_path: str, place: str) -> str:
    """
    Entry point for Chaquopy:
    - Reads metadata.json
    - Processes input Excel at input_path
    - Writes output to /sdcard/Documents/bills
    - Returns output file path
    """
    # Load metadata
    meta = json.loads(METADATA_FILE.read_text())
    if place not in meta or place in ("GST", "vendor_code"):
        # Fallback to first real key
        place = next(k for k in meta if k not in ("GST", "vendor_code"))

    po, delivery_date = _extract_po_and_date(Path(input_path).name)
    df    = pd.read_excel(input_path)
    items = _transform_items(df)

    invoice_no = str(_get_next_invoice_number())
    pm = meta[place]

    bill_data = {
        "GST": meta["GST"],
        "vendor_code": meta["vendor_code"],
        "PO": po,
        "delivery_date": delivery_date,
        "invoice_no": invoice_no,
        "bill_to": pm["bill_to"],
        "place_of_supply": pm["place_of_supply"],
        "site_code": pm["site_code"],
        "items": items,
    }

    # Prepare output folder and filename
    out_dir = Path("/sdcard/Documents/bills")
    out_dir.mkdir(parents=True, exist_ok=True)
    date_part = datetime.strptime(delivery_date, "%d-%m-%Y").strftime("%Y-%m-%d")
    out_file = f"{place}_{date_part}_{po}.xlsx"
    out_path = out_dir / out_file

    # Build the Excel
    _create_excel(bill_data, out_path)
    pdf_path = str(out_path.with_suffix(".pdf"))
    generate_pdf_bill(bill_data, pdf_path)
    return json.dumps({"excel": str(out_path), "pdf": pdf_path})

def _create_excel(data: dict, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # — HEADER —
    title_texts = ["TAX INVOICE", "A.G AGRO", "TEGHRA,BEGUSARAI-851133", data["GST"]]
    for i, text in enumerate(title_texts, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(size=(15 if i==1 else 12), bold=True, italic=True)
        ws.row_dimensions[i].height = (36 if i==1 else 14)

    # — BILL INFO —
    ws.merge_cells("A5:D5")
    ws.merge_cells("E5:H5")
    ws.merge_cells("I5:L5")
    
    section_headers = {"A5": "BILL TO", "E5": "PLACE OF SUPPLY", "I5": "BILL DETAILS:"}

    for cell_ref, text in section_headers.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.font = Font(bold=True, italic=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 12

    ws.merge_cells("A6:D7"); ws["A6"].value = data["bill_to"]
    ws.merge_cells("E6:H7"); ws["E6"].value = data["place_of_supply"]
    ws.merge_cells("I6:L7")
    ws["I6"].value = (
        f"INVOICE NO: {data['invoice_no']}\n"
        f"DATE: {data['delivery_date']}\n"
        f"VENDOR: {data['vendor_code']}\n"
        f"SITE CODE: {data['site_code']}"
    )
    for col in range(1, 13):
        for row in range(6, 8):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for cell in ["A6", "E6", "I6"]:
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cell].font = Font(size=9, bold=True, italic=True)

    for row in range(6, 8):
        ws.row_dimensions[row].height = 36
        
    # === GST and PO ===
    ws.merge_cells("A8:F8")
    ws.merge_cells("G8:L8")
    ws["A8"] = f"GST: {data['GST']}"
    ws["G8"] = f"PO-{data['PO']}"
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A8"].font = Font(bold=True, italic=True)
    ws["G8"].font = Font(bold=True, italic=True)
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws["A8"].fill = blue_fill
    ws["G8"].fill = blue_fill

    ws.row_dimensions[8].height = 12

    # — TABLE HEADERS —
    headers = [
        "ARTICLE CODE", "HSN CODE", "Article Description", "Grammage", "Quantity",
        "Rate", "Taxable Value", "SGST Rate", "SGST Amount",
        "CGST Rate", "CGST Amount", "Total Amount"
    ]
    hr = 9
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=idx, value=h)
        cell.fill      = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font      = Font(bold=True, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(*[Side("thin")]*4)
    ws.row_dimensions[hr].height = 30

    # — DATA ROWS —
    start = hr + 1
    for r, item in enumerate(data["items"]):
        for c, val in enumerate(item, start=1):
            cell = ws.cell(row=start + r, column=c, value=val)
            cell.alignment = Alignment(wrapText=True, vertical="top")
            if c in (1,2):
                cell.number_format = '0'
            elif c in (5,6,7,9,11,12):
                cell.number_format = '#,##0.00'
            elif c in (8,10):
                cell.number_format = '0.00%'
            cell.border = Border(*[Side("thin")]*4)
            if r%2==0:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")

    # — FOOTER SUMS —
    end = start + len(data["items"])
    ws.cell(row=end, column=1, value="Sub Total").alignment = Alignment(horizontal="right", vertical="center")
    for col in (5,7,8,9,10,11,12):
        letter = get_column_letter(col)
        cell = ws.cell(
            row=end,
            column=col,
            value=f"=SUM({letter}{start}:{letter}{end-1})"
        )
        cell.number_format = '#,##0.00' if col not in (8,10) else '0.00%'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = Border(*[Side("thin")]*4)

    min_width = 10
    max_width = 40
    max_row = end + 1

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col - 1])
        for row in range(start, max_row):
            value = ws.cell(row=row, column=col).value
            if value:
                max_length = max(max_length, len(str(value)))
        adjusted_width = min(max(max_length, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

        # === ADDITIONAL FOOTER INFO ===
    current_row = end + 1

    # Row: Misc. Charges
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    misc_cell = ws.cell(row=current_row, column=1, value="Misc. Charges (Including Freight, Octroi, Loading, Unloading, etc.)")
    misc_cell.font = Font(bold=True, size=9, italic=True)
    misc_cell.alignment = Alignment(horizontal="right", vertical="center")
    current_row += 1

    # Row: Grand Total
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    grand_cell = ws.cell(row=current_row, column=1, value="Grand Total (Rounded Off)")
    grand_cell.font = Font(bold=True, size=9 ,italic=True)
    grand_cell.alignment = Alignment(horizontal="right", vertical="center")
    current_row += 1

    # Blank Row
    current_row += 1

    # Thank You Row (Center-aligned)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    thank_you_cell = ws.cell(row=current_row, column=1, value="THANK YOU FOR YOUR BUSINESS")
    thank_you_cell.font = Font(bold=True, size=8, italic=True)
    thank_you_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Clear borders for entire "THANK YOU" row (including the merged cell)
    no_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None),
    )

    for col in range(1, 13):
        cell = ws.cell(row=current_row, column=col)
        cell.border = no_border

    current_row += 1

    # Clear borders for 3 rows after "THANK YOU"
    for row in range(current_row, current_row + 3):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = no_border

    # Signature at bottom-left
    signature_cell = ws.cell(row=current_row + 3, column=1, value="Signature")
    signature_cell.font = Font(bold=True, size=8)
    signature_cell.alignment = Alignment(horizontal="left", vertical="center")

    # — PRINT PREVIEW: A4 landscape, fit to width, custom margins —
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # — SAVE —
    wb.save(out_path)
